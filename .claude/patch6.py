import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from collections import defaultdict
from datetime import datetime, timedelta

wb = openpyxl.load_workbook(
    r'C:\Users\mkyle\OneDrive\Desktop\darp dashboard\data files\S2026 Data Analysis .xlsx',
    data_only=True
)
ws = wb['Bulk Data']

# col A=timestamp, B=name, E=hours(timedelta), H=lbs, I=lb/hr, M=hall(MF/BF), N=Efficiency Score
# Print first 5 rows to understand timestamp format and available columns
print("=== COLUMN HEADERS ===")
headers = [ws.cell(1, c).value for c in range(1, 20)]
print(headers)

print("\n=== SAMPLE ROWS (first 5 data rows) ===")
for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):
    print(row[:15])
